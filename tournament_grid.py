import math
import os

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font


def create_grid(users, cat_name, comp_name, city, members, tyli=False):
    for i, name in enumerate(users):
        users[i] = f'{i + 1}.' + name
    first_pow2 = 2 ** math.ceil(math.log2(len(users)))
    first_byes = first_pow2 - len(users)
    first_round_players = len(users) - first_byes
    rounds_count = math.ceil(math.log2(len(users)))

    filename = f'Тыли {comp_name}.xlsx' if tyli else f'Массоги {comp_name}.xlsx'
    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.create_sheet(title=cat_name.replace('Массоги', '').replace('Тыли', '').replace(' ', ''))
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = cat_name.replace('Массоги', '').replace('Тыли', '').replace(' ', '')

    bot_border = Border(bottom=Side(style="thin"))
    right_border = Border(right=Side(style="thin"))
    bot_right_border = Border(right=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    bold = Font(bold=True, size=14)
    sheet["A1"] = comp_name
    sheet["A2"] = city
    sheet["A3"] = cat_name
    for i in ['A1', 'A2', 'A3']:
        sheet[i].alignment = center
        sheet[i].font = bold
    sheet.merge_cells("A1:D1")
    sheet.merge_cells("A2:D2")
    sheet.merge_cells("A3:D3")
    for i in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        sheet.column_dimensions[i].width = 50

    row_index = 5
    current_round = 0
    last_column = 1
    for i, name in enumerate(users):
        if first_byes and name in users[-first_byes:]:
            row_index += 1
            cell = sheet.cell(row=row_index, column=2, value=name)
            row_index += 3
            continue
        else:
            cell = sheet.cell(row=row_index, column=1, value=name)

        cell.border = bot_border
        row_index += 2
        if i % 2 == 0:
            cell = sheet.cell(row=row_index - 1, column=1)
            cell.border = right_border
        else:
            cell.border = bot_right_border

    current_round += 1
    row_index = 6
    next_last_row_counter = 2
    last_row = (first_round_players // 2 + first_byes) * 4 - next_last_row_counter + 4
    if rounds_count >= current_round:
        last_column = 2
        while row_index <= last_row:
            if rounds_count == current_round:
                cell = sheet.cell(row=row_index + 1, column=2, value='1 место')
            cell = sheet.cell(row=row_index, column=2)
            if row_index in [i for i in range(10, 800, 8)]:
                cell.border = bot_right_border
            elif row_index in [i for i in range(6, 800, 8)]:
                cell.border = bot_border
            else:
                cell.border = right_border
            row_index += 1
            if row_index in [i for i in range(11, 800, 8)]:
                row_index += 3

        current_round += 1
        row_index = 8
        last_row -= next_last_row_counter

        if rounds_count >= current_round:
            last_column = 3
            while row_index <= last_row:
                if rounds_count == current_round:
                    cell = sheet.cell(row=row_index + 1, column=3, value='1 место')
                cell = sheet.cell(row=row_index, column=3)
                if row_index in [i for i in range(16, 800, 16)]:
                    cell.border = bot_right_border
                elif row_index in [i for i in range(8, 800, 16)]:
                    cell.border = bot_border
                else:
                    cell.border = right_border
                row_index += 1
                if row_index in [i for i in range(17, 800, 16)]:
                    row_index += 7

            current_round += 1
            row_index = 12
            next_last_row_counter *= 2
            last_row -= next_last_row_counter

            if rounds_count >= current_round:
                last_column = 4
                while row_index <= last_row:
                    if rounds_count == current_round:
                        cell = sheet.cell(row=row_index + 1, column=4, value='1 место')
                    cell = sheet.cell(row=row_index, column=4)
                    if row_index in [i for i in range(28, 800, 32)]:
                        cell.border = bot_right_border
                    elif row_index in [i for i in range(12, 800, 32)]:
                        cell.border = bot_border
                    else:
                        cell.border = right_border
                    row_index += 1
                    if row_index in [i for i in range(29, 800, 32)]:
                        row_index += 15

                current_round += 1
                row_index = 20
                next_last_row_counter *= 2
                last_row -= next_last_row_counter
                if rounds_count >= current_round:
                    last_column = 5
                    while row_index <= last_row:
                        if rounds_count == current_round:
                            cell = sheet.cell(row=row_index + 1, column=5, value='1 место')
                        cell = sheet.cell(row=row_index, column=5)
                        if row_index in [i for i in range(52, 800, 64)]:
                            cell.border = bot_right_border
                        elif row_index in [i for i in range(20, 800, 64)]:
                            cell.border = bot_border
                        else:
                            cell.border = right_border
                        row_index += 1
                        if row_index in [i for i in range(53, 800, 64)]:
                            row_index += 31

                    current_round += 1
                    row_index = 36
                    next_last_row_counter *= 2
                    last_row -= next_last_row_counter
                    if rounds_count >= current_round:
                        last_column = 6
                        while row_index <= last_row:
                            if rounds_count == current_round:
                                cell = sheet.cell(row=row_index + 1, column=6, value='1 место')
                            cell = sheet.cell(row=row_index, column=6)
                            if row_index in [i for i in range(100, 800, 128)]:
                                cell.border = bot_right_border
                            elif row_index in [i for i in range(36, 800, 128)]:
                                cell.border = bot_border
                            else:
                                cell.border = right_border
                            row_index += 1
                            if row_index in [i for i in range(101, 800, 128)]:
                                row_index += 63

    cell = sheet.cell(row=sheet.max_row + 1, column=last_column - 1)
    cell.border = bot_border
    cell = sheet.cell(row=sheet.max_row + 1, column=last_column - 1)
    cell.border = right_border
    cell = sheet.cell(row=sheet.max_row, column=last_column)
    cell.border = bot_border
    cell = sheet.cell(row=sheet.max_row + 1, column=last_column - 1)
    cell.border = bot_right_border
    cell = sheet.cell(row=sheet.max_row, column=last_column, value='3 место')
    cell.alignment = center

    last_row = sheet.max_row + 2
    winners = ['1 место:', '2 место:', '3 место:']
    for i, text in enumerate(members.keys()):
        cell1 = sheet.cell(row=last_row, column=last_column - 1, value=text)
        cell = sheet.cell(row=last_row, column=1, value=winners[i])
        cell.border = bot_border
        cell2 = sheet.cell(row=last_row, column=last_column, value=f'/{members[text]}')
        last_row += 2
        cell2.border = bot_border
        cell2.alignment = right

    workbook.save(filename)

