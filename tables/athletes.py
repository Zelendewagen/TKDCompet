import os
import sqlite3
import traceback
from datetime import datetime
from tkinter import ttk, messagebox
import tkinter as tk
from tkinter.filedialog import askopenfilename

import pandas as pd
from openpyxl.styles import Border, Side
from openpyxl.workbook import Workbook

import config
from config import DB_FILE


class AthletesFrame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.current_id = None
        self.create_top_buttons()
        self.create_table()
        self.categories = [f"{i} гып" for i in range(10, 0, -1)] +[f"{i} дан" for i in range(1, 10)]

    def create_top_buttons(self):
        buttons_frame = tk.Frame(self)
        back_button = ttk.Button(buttons_frame, text="\u2190", command=self.show_competitions_table, width=10)
        add_button = ttk.Button(buttons_frame, text="Добавить участника", command=self.open_add_change_athlete_window,
                                width=20)
        add_list_button = ttk.Button(buttons_frame, text="Добавить список", command=self.add_list_athletes, width=20)
        clear_list_button = ttk.Button(buttons_frame, text="Удалить всех", command=self.delete_all_athletes, width=20)
        save_list = ttk.Button(buttons_frame, text="Сохранить таблицу", command=self.save_xls,
                               width=20)
        create_tyli = ttk.Button(buttons_frame, text="Тыли", command=self.show_tyli, width=20)
        create_massogi = ttk.Button(buttons_frame, text="Массоги", command=self.show_massogi, width=20)

        back_button.pack(side="left")
        add_button.pack(side="left", padx=5)
        add_list_button.pack(side="left", padx=5)
        clear_list_button.pack(side="left", padx=5)
        save_list.pack(side="left", padx=5)
        create_tyli.pack(side="left", padx=5)
        create_massogi.pack(side="left", padx=5)

        buttons_frame.pack(fill="both", pady=(0, 10))

    def create_table(self):
        heads = ['№', 'ФИО', 'Пол', 'Дата рождения', 'Полных лет', 'Вес', 'Категория', 'Город', 'Клуб', 'Тренер',
                 'Тыли', 'Массоги', 'ID']
        self.table = ttk.Treeview(self, columns=heads, show='headings', selectmode="browse")
        for header in heads:
            self.table.heading(header, text=header, anchor='w')
        self.table.column('№', stretch=False, minwidth=50, width=50)
        self.table.column('ФИО', minwidth=150)
        self.table.column('Пол', stretch=False, minwidth=80, width=80)
        self.table.column('Дата рождения', stretch=False, minwidth=80, width=100)
        self.table.column('Полных лет', stretch=False, minwidth=80, width=80)
        self.table.column('Вес', stretch=False, minwidth=50, width=50)
        self.table.column('Категория', stretch=False, minwidth=80, width=80)
        self.table.column('Город', stretch=True, minwidth=80, width=80)
        self.table.column('Клуб', stretch=True, minwidth=80, width=80)
        self.table.column('Тренер', stretch=True, minwidth=80, width=80)
        self.table.column('Тыли', stretch=True, minwidth=80, width=80)
        self.table.column('Массоги', stretch=True, minwidth=80, width=80)
        self.table.column('ID', stretch=False, minwidth=50, width=50)
        self.table.bind('<Button-3>', self.right_button_menu)
        self.table.bind("<Double-1>", self.on_double_click)

        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.table.yview)
        self.table.configure(yscroll=scrollbar.set)
        self.table.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def right_button_menu(self, event):
        item = self.table.identify_row(event.y)
        if item:
            self.table.selection_set(item)
            self.table.focus(item)
            values = self.table.item(item, "values")

        event.widget.focus()
        file_menu = tk.Menu(self)
        file_menu.add_command(label='Изменить',
                              command=lambda: self.open_add_change_athlete_window(change=True, athlete_id=values[-1]))
        file_menu.add_separator()
        file_menu.add_command(label='Удалить', command=lambda: self.delete_athlete(values[-1]))
        file_menu.post(event.x_root, event.y_root)

    def on_double_click(self, event):
        item = self.table.identify_row(event.y)
        if item:
            values = self.table.item(item, "values")
            self.open_add_change_athlete_window(change=True, athlete_id=values[-1])

    def open_add_change_athlete_window(self, change=False, athlete_id=None):
        self.top_window = tk.Toplevel()
        self.top_window.grab_set()
        title = 'Изменить участника' if change else 'Добавить участника'
        self.top_window.title(title)
        self.top_window.geometry(config.geometry(self.master, config.athl_window_width, config.athl_window_height))
        self.top_window.resizable(False, False)
        self.top_window.transient(self)

        frame = tk.Frame(self.top_window)
        frame.pack(fill="both", padx=(10, 50), pady=10, expand=True)
        frame.columnconfigure(index=1, weight=250)

        labels = ['ФИО', 'Пол', 'Дата рождения', 'Вес', 'Категория', 'Город', 'Клуб', 'Тренер', 'Тыли', 'Массоги']
        for i, text in enumerate(labels):
            label = tk.Label(frame, text=text)
            if i == len(labels) - 1:
                label.grid(column=0, row=i, sticky='w')
            else:
                label.grid(column=0, row=i, sticky='w', pady=(0, 10))

        entries = [tk.Entry(frame, width=30) for _ in range(len(labels))]
        entries[1] = ttk.Combobox(frame, values=['муж', 'жен'], state="readonly", width=30)
        entries[3] = ttk.Combobox(frame, values=[str(i) for i in range(1, 150)], state="readonly", width=30)
        entries[4] = ttk.Combobox(frame, values=self.categories, state="readonly", width=30)
        entries[8] = ttk.Combobox(frame, values=['да', 'нет'], state="readonly", width=30)
        entries[9] = ttk.Combobox(frame, values=['да', 'нет'], state="readonly", width=30)
        for i, entry in enumerate(entries):
            if i + 1 == len(entries):
                entries[i].grid(column=1, row=i, sticky='ew')
            else:
                entries[i].grid(column=1, row=i, sticky='ew', pady=(0, 10))

        buttons_frame = tk.Frame(frame)
        buttons_frame.grid(columnspan=2, row=len(labels), pady=20, padx=(20, 0), sticky="ew")

        if change:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM athletes WHERE id =?", (athlete_id,))
                values = cursor.fetchone()
            entries[1].set(values[3])
            entries[3].set(values[6])
            entries[4].set(self.categories[values[7]])
            entries[8].set(values[11])
            entries[9].set(values[12])
            for i, entry in enumerate(entries):
                if i >= 3:
                    entries[i].insert(0, values[i + 3])
                else:
                    entries[i].insert(0, values[i + 2])

            change_button = ttk.Button(buttons_frame, text='Изменить', width=15,
                                       command=lambda:
                                       self.add_change_athlete(entries, athlete_id=values[0], change=True))
            change_button.pack(fill="both", expand=True, side="left", padx=5)
        else:
            add_button = ttk.Button(buttons_frame, text='Добавить', width=15,
                                    command=lambda: self.add_change_athlete(entries))
            add_button.pack(fill="both", expand=True, side="left", padx=5)

        cancel_button2 = ttk.Button(buttons_frame, text="Отмена", command=self.top_window.destroy, width=15)
        cancel_button2.pack(fill="both", expand=True, side="left", padx=5)

    def add_list_athletes(self):
        file = askopenfilename(title='Открыть', filetypes=(("Таблица Excel", "*.xls"),))
        if file:
            xls = pd.ExcelFile(file)
            sheet_names = xls.sheet_names
            data = []
            sheets = []
            current_row = 0
            for sheet in sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet)
                if not df.dropna(how="all").empty:
                    data = df.values.tolist()
                    sheets.append(sheet)
                    data = data[3:]
            if data:
                if len(sheets) == 1:
                    try:
                        with sqlite3.connect(DB_FILE) as conn:
                            cursor = conn.cursor()
                            for row in data:
                                current_row = data.index(row) + 1
                                cursor.execute("SELECT date FROM competitions WHERE id = ?", (self.current_id,))
                                birth_date = datetime.strptime(row[3].strftime("%d.%m.%Y"), "%d.%m.%Y")
                                compet_date = datetime.strptime(cursor.fetchone()[0], "%d.%m.%Y")
                                age = compet_date.year - birth_date.year
                                if (compet_date.month, compet_date.day) < (birth_date.month, birth_date.day):
                                    age -= 1
                                parameters = (
                                    self.current_id, row[2], row[1], row[3].strftime("%d.%m.%Y"), age,
                                    int(float(row[4])),
                                    self.categories.index(row[6]), row[8], row[11], row[12], row[13], row[14])
                                cursor.execute("""INSERT INTO athletes (competition_id, name, gender, date, age, weight, category, location, club, trainer, tyli, massogi) 
                                                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                                        """, parameters)
                            conn.commit()
                        self.load_table(self.current_id)
                        messagebox.showinfo("Готово", f"Участники загружены из {os.path.basename(file)} {sheets[0]}")
                    except Exception as e:
                        messagebox.showerror("Ошибка add_list", f"Возможно ошибка в строке: {current_row}!\n\n{e}",
                                             parent=self.master)
                        messagebox.showerror("Ошибка", traceback.format_exc())
                        raise
                else:
                    messagebox.showerror("Ошибка", f"Загрузка отменена, проверьте {', '.join(map(str, sheets))}")

    def add_change_athlete(self, entries, athlete_id=None, change=False):
        parameters = ()
        for i, entry in enumerate(entries):
            if i == 4:
                parameters += (self.categories.index(entry.get()),)
            else:
                parameters += (entry.get(),)
        try:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                datetime.strptime(parameters[2], "%d.%m.%Y")

                cursor.execute("SELECT date FROM competitions WHERE id = ?", (self.current_id,))
                birth_date = datetime.strptime(parameters[2], "%d.%m.%Y")
                compet_date = datetime.strptime(cursor.fetchone()[0], "%d.%m.%Y")
                age = compet_date.year - birth_date.year
                if (compet_date.month, compet_date.day) < (birth_date.month, birth_date.day):
                    age -= 1
                parameters += (age,)
                if change:
                    parameters += (athlete_id,)
                    cursor.execute(
                        "UPDATE athletes SET name = ?, gender = ?, date = ?, weight = ?, category = ?, location = ?, club = ?, trainer = ?, tyli = ?, massogi = ?, age = ? WHERE id = ?",
                        parameters)
                else:
                    parameters += (self.current_id,)
                    cursor.execute("""INSERT INTO athletes (name, gender, date, weight, category, location, club, trainer, tyli, massogi, age, competition_id)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, parameters)
                conn.commit()
            self.load_table(self.current_id)
            self.top_window.destroy()
        except ValueError:
            messagebox.showerror("Ошибка", "Не верный формат даты!", parent=self.master)

    def delete_athlete(self, athlete_id):
        answer = messagebox.askyesno("Удалить участника", "УДАЛИТЬ?")
        if answer:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM athletes WHERE id = ?", (athlete_id,))
                conn.commit()
            self.load_table(self.current_id)

    def delete_all_athletes(self):
        answer = messagebox.askyesno("Удалить всех участников", "Удалить ВСЕХ участников?")
        if answer:
            with sqlite3.connect(DB_FILE) as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM athletes WHERE competition_id = ?", (self.current_id,))
                conn.commit()
            self.load_table(self.current_id)

    def load_table(self, comp_id):
        self.current_id = comp_id
        for row in self.table.get_children():
            self.table.delete(row)
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM athletes WHERE competition_id = ?", (self.current_id,))
            athletes = cursor.fetchall()

        for num, row in enumerate(athletes):
            data = [i for i in row[2:]] + [row[0]]
            data[5] = self.categories[data[5]]
            values = [num + 1] + data
            self.table.insert("", tk.END, values=values)

    def show_competitions_table(self):
        self.master.show_table('соревнования', self)

    def show_tyli(self):
        self.master.show_table('тыли', self)
        self.master.tyli_frame.update_all_users(self.current_id)
        self.master.tyli_frame.update_users_category()
        self.master.tyli_frame.update_tables()

    def show_massogi(self):
        self.master.show_table('массоги', self)
        self.master.massogi_frame.update_all_users(self.current_id)
        self.master.massogi_frame.update_users_category()
        self.master.massogi_frame.update_tables()

    @staticmethod
    def update_ages(comp_id):
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, date FROM athletes WHERE competition_id = ?", (comp_id,))
            rows = cursor.fetchall()
            cursor.execute("SELECT date FROM competitions WHERE id = ?", (comp_id,))
            compet_date = datetime.strptime(cursor.fetchone()[0], "%d.%m.%Y")
            for ath_id, date in rows:
                birth_date = datetime.strptime(date, "%d.%m.%Y")
                age = compet_date.year - birth_date.year
                if (compet_date.month, compet_date.day) < (birth_date.month, birth_date.day):
                    age -= 1
                cursor.execute("UPDATE athletes SET age = ? WHERE id = ?", (age, ath_id))

    def save_xls(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "заявка"

        headers = ["ФИО", "дата рождения", "Пол", "Вес", "Квалификация", "Клуб", "Тренер", "Возраст"]
        ws.append(headers)
        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM athletes WHERE competition_id = ?", (self.current_id,))
            for row in cursor.fetchall():
                ws.append([row[2], row[4], row[3], row[6], self.categories[row[7]], row[9], row[10], row[5]])

        thick_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = thick_border

        with sqlite3.connect(DB_FILE) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM competitions WHERE id = ?", (self.current_id,))
            name = cursor.fetchone()[0]

        wb.save(f"{name}.xlsx")
        messagebox.showinfo("TKD", f"Таблица сохранена: {name}.xlsx")
